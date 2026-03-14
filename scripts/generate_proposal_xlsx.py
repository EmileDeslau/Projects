#!/usr/bin/env python3
"""Generate the professional Excel pricing proposal for the restaurant project.
Usage: python scripts/generate_proposal_xlsx.py [--rate 300] [--tva 20]
"""
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from pathlib import Path


def make_styles():
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return {
        "title": Font(name="Calibri", size=18, bold=True, color="1F4E79"),
        "subtitle": Font(name="Calibri", size=12, italic=True, color="555555"),
        "header": Font(name="Calibri", size=11, bold=True, color="FFFFFF"),
        "header_fill": PatternFill("solid", fgColor="1F4E79"),
        "module_font": Font(name="Calibri", size=11, bold=True, color="1F4E79"),
        "module_fill": PatternFill("solid", fgColor="D6E4F0"),
        "total_font": Font(name="Calibri", size=11, bold=True),
        "total_fill": PatternFill("solid", fgColor="E2EFDA"),
        "grand_total_font": Font(name="Calibri", size=12, bold=True, color="FFFFFF"),
        "grand_total_fill": PatternFill("solid", fgColor="1F4E79"),
        "border": border,
        "wrap": Alignment(wrap_text=True, vertical="center"),
        "center": Alignment(horizontal="center", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
        "price_fmt": '#,##0 €',
    }


FEATURES = [
    # (ref, name, detail, days, phase, module)
    # Phase 1
    ("M-01", "Création / modification / suppression d'articles", "CRUD complet : nom, description, catégorie, image, statut", 3, 1, "Gestion du Menu"),
    ("M-02", "Paliers de prix par quantité", "Interface de configuration des tranches tarifaires par article", 5, 1, "Gestion du Menu"),
    ("M-03", "Gestion des catégories", "Entrées, Plats, Desserts, Boissons, Formules", 1.5, 1, "Gestion du Menu"),
    ("M-04", "Variantes d'article", "Taille, cuisson, suppléments avec tarification propre", 3, 1, "Gestion du Menu"),
    ("M-05", "Activation / désactivation d'article", "Masquer un plat en rupture de stock", 0.5, 1, "Gestion du Menu"),
    ("M-06", "Affichage du prix dynamique", "Recalcul en temps réel selon la quantité", 4, 1, "Gestion du Menu"),
    ("M-07", "Créer une commande", "Association table ou emporter/livraison", 2, 1, "Commandes"),
    ("M-08", "Ajouter articles avec recalcul prix", "Sélection, quantité, recalcul auto par palier", 4, 1, "Commandes"),
    ("M-09", "Modifier une commande en cours", "Ajout/retrait d'articles, changement de quantités", 2, 1, "Commandes"),
    ("M-10", "Annuler une commande", "Annulation totale/partielle avec motif", 1.5, 1, "Commandes"),
    ("M-11", "Workflow de statuts", "En attente → Préparation → Prête → Servie → Clôturée", 2, 1, "Commandes"),
    ("M-12", "Notes et allergies", "Instructions spéciales, tags allergènes", 1, 1, "Commandes"),
    ("M-13", "Plan de salle configurable", "Placement des tables, capacité, vue graphique", 4, 1, "Tables"),
    ("M-14", "Statuts des tables", "Libre / Occupée / Réservée / À nettoyer", 1.5, 1, "Tables"),
    ("M-15", "Lien table ↔ commande", "Association automatique et vue depuis le plan", 1.5, 1, "Tables"),
    ("M-16", "Génération de l'addition", "Détail par article avec palier, TVA, total", 3, 1, "Facturation"),
    ("M-17", "Modes de paiement", "CB, Espèces, Ticket restaurant, paiement mixte", 2, 1, "Facturation"),
    ("M-18", "Division de l'addition", "Split par convive ou par article", 2, 1, "Facturation"),
    ("M-19", "Impression / envoi ticket", "Imprimante thermique, e-mail, PDF", 2.5, 1, "Facturation"),
    ("M-20", "TVA configurable", "Taux par catégorie, sur place vs emporter", 1.5, 1, "Facturation"),
    # Phase 2
    ("M-21", "Écran de commandes cuisine (KDS)", "Affichage temps réel, tri, code couleur", 4, 2, "Cuisine (KDS)"),
    ("M-22", "Marquage « Prêt »", "Validation cuisinier, notification serveur", 2, 2, "Cuisine (KDS)"),
    ("M-23", "Connexion sécurisée", "Login/mot de passe + code PIN rapide", 3, 2, "Authentification"),
    ("M-24", "Rôles et permissions", "Admin, Serveur, Cuisinier, Caissier", 3, 2, "Authentification"),
    ("M-25", "Journal d'activité", "Audit log complet, non modifiable", 2, 2, "Authentification"),
    # Phase 3
    ("S-01", "Menu en ligne + QR code", "Page responsive avec prix dynamiques", 4, 3, "Commande en ligne"),
    ("S-02", "Commande emporter / livraison", "Panier interactif, prix temps réel", 5, 3, "Commande en ligne"),
    ("S-03", "Paiement en ligne", "Intégration Stripe / PayPal", 4, 3, "Commande en ligne"),
    ("S-04", "Suivi de commande client", "Page de suivi temps réel + notifications", 3, 3, "Commande en ligne"),
    # Phase 4
    ("S-05", "Suivi des ingrédients", "CRUD ingrédients, quantités, seuils", 3, 4, "Stocks"),
    ("S-06", "Déduction automatique", "Réduction stock à chaque commande", 2.5, 4, "Stocks"),
    ("S-07", "Alertes de rupture", "Notifications + désactivation auto", 1.5, 4, "Stocks"),
    ("S-08", "Système de réservation", "Date, heure, couverts, conflits", 3, 4, "Réservations"),
    ("S-09", "Confirmation automatique", "E-mail/SMS confirmation + rappel J-1", 2, 4, "Réservations"),
    ("S-10", "Intégration plan de salle", "Attribution automatique des tables", 2, 4, "Réservations"),
    ("S-11", "Compte client", "Historique, préférences", 2, 4, "Fidélité"),
    ("S-12", "Points de fidélité", "Cumul automatique selon le montant", 2, 4, "Fidélité"),
    ("S-13", "Réductions par paliers", "Récompenses configurables", 1.5, 4, "Fidélité"),
    ("S-14", "Dashboard temps réel", "CA, commandes, taux d'occupation", 3, 4, "Rapports"),
    ("S-15", "Rapports de vente", "Par période, article, catégorie", 3, 4, "Rapports"),
    ("S-16", "Analyse des paliers de prix", "Impact des tarifs dégressifs", 2, 4, "Rapports"),
    ("S-17", "Export CSV / PDF", "Pour la comptabilité", 1.5, 4, "Rapports"),
    ("S-18", "Interface multilingue", "Français / Anglais minimum", 2, 4, "UX"),
    ("S-19", "Accessibilité WCAG 2.1", "Contraste, clavier, lecteur d'écran", 2, 4, "UX"),
]

PHASE_NAMES = {
    1: "Phase 1 — MVP",
    2: "Phase 2 — Cuisine & Rôles",
    3: "Phase 3 — Commande en ligne",
    4: "Phase 4 — Avancé",
}


def build_proposal_sheet(wb, rate, tva_pct):
    ws = wb.active
    ws.title = "Proposition Commerciale"
    s = make_styles()

    # Column widths
    widths = {"A": 8, "B": 36, "C": 42, "D": 10, "E": 14}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "Proposition Commerciale — Application Restaurant"
    c.font = s["title"]
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:E2")
    c2 = ws["A2"]
    c2.value = f"Taux journalier : {rate:.0f} € HT/jour  |  TVA : {tva_pct:.0f}%  |  Date : 14/03/2026"
    c2.font = s["subtitle"]
    c2.alignment = Alignment(horizontal="center")

    # Headers row 4
    headers = ["Réf.", "Fonctionnalité", "Détail", "Jours", "Prix HT"]
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.font = s["header"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["border"]

    row = 5
    phase_totals = {}
    current_module = None

    for ref, name, detail, days, phase, module in FEATURES:
        # Phase header
        phase_label = PHASE_NAMES[phase]
        if phase not in phase_totals:
            if phase > 1:
                row += 1  # blank row between phases
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=1, value=phase_label)
            cell.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="2E75B6")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 28
            row += 1
            phase_totals[phase] = {"days": 0, "price": 0, "start_row": row}
            current_module = None

        # Module sub-header
        if module != current_module:
            current_module = module
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            cell = ws.cell(row=row, column=1, value=f"  {module}")
            cell.font = s["module_font"]
            cell.fill = s["module_fill"]
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = s["border"]
            row += 1

        # Feature row
        price = max(days * rate, 150)
        price = round(price / 25) * 25  # round to nearest 25

        ws.cell(row=row, column=1, value=ref).alignment = s["center"]
        ws.cell(row=row, column=2, value=name).alignment = s["wrap"]
        ws.cell(row=row, column=3, value=detail).alignment = s["wrap"]
        d_cell = ws.cell(row=row, column=4, value=days)
        d_cell.alignment = s["center"]
        d_cell.number_format = "0.0"
        p_cell = ws.cell(row=row, column=5, value=price)
        p_cell.alignment = s["right"]
        p_cell.number_format = s["price_fmt"]
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = s["border"]

        phase_totals[phase]["days"] += days
        phase_totals[phase]["price"] += price
        row += 1

        # Check if last feature of this phase
        next_features = [f for f in FEATURES if f[4] == phase]
        is_last = (ref == next_features[-1][0])
        if is_last:
            # Phase subtotal
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            cell = ws.cell(row=row, column=1, value=f"TOTAL {phase_label}")
            cell.font = s["total_font"]
            cell.fill = s["total_fill"]
            cell.alignment = Alignment(horizontal="right", vertical="center")
            d = ws.cell(row=row, column=4, value=phase_totals[phase]["days"])
            d.font = s["total_font"]
            d.fill = s["total_fill"]
            d.alignment = s["center"]
            d.number_format = "0.0"
            p = ws.cell(row=row, column=5, value=phase_totals[phase]["price"])
            p.font = s["total_font"]
            p.fill = s["total_fill"]
            p.alignment = s["right"]
            p.number_format = s["price_fmt"]
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = s["border"]
            row += 1

    # Grand totals
    row += 1
    total_days = sum(v["days"] for v in phase_totals.values())
    total_ht = sum(v["price"] for v in phase_totals.values())
    total_tva = total_ht * tva_pct / 100
    total_ttc = total_ht + total_tva

    for label, val, fmt in [
        ("TOTAL PROJET HT", total_ht, s["price_fmt"]),
        (f"TVA ({tva_pct:.0f}%)", total_tva, s["price_fmt"]),
        ("TOTAL PROJET TTC", total_ttc, s["price_fmt"]),
    ]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        cell = ws.cell(row=row, column=1, value=label)
        cell.font = s["grand_total_font"]
        cell.fill = s["grand_total_fill"]
        cell.alignment = Alignment(horizontal="right", vertical="center")
        p = ws.cell(row=row, column=5, value=val)
        p.font = s["grand_total_font"]
        p.fill = s["grand_total_fill"]
        p.alignment = s["right"]
        p.number_format = fmt
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = s["border"]
        ws.row_dimensions[row].height = 30
        row += 1

    # Options section
    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws["A" + str(row)].value = "Options de Livraison"
    ws["A" + str(row)].font = s["title"]
    row += 1

    options_headers = ["Option", "Phases incluses", "", "Montant HT", "Montant TTC"]
    for i, h in enumerate(options_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = s["header"]
        cell.fill = s["header_fill"]
        cell.alignment = s["center"]
        cell.border = s["border"]
    row += 1

    p1 = phase_totals[1]["price"]
    p2 = phase_totals[2]["price"]
    p3 = phase_totals[3]["price"]
    options = [
        ("Essentiel", "Phase 1", p1),
        ("Standard", "Phase 1 + 2", p1 + p2),
        ("Complet", "Phase 1 + 2 + 3", p1 + p2 + p3),
        ("Premium", "Toutes les phases", total_ht),
    ]
    for opt_name, phases, ht in options:
        ws.cell(row=row, column=1, value=opt_name).font = Font(bold=True)
        ws.cell(row=row, column=1).alignment = s["center"]
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws.cell(row=row, column=2, value=phases).alignment = s["center"]
        ws.cell(row=row, column=4, value=ht).number_format = s["price_fmt"]
        ws.cell(row=row, column=4).alignment = s["right"]
        ws.cell(row=row, column=5, value=ht * (1 + tva_pct / 100)).number_format = s["price_fmt"]
        ws.cell(row=row, column=5).alignment = s["right"]
        for c in range(1, 6):
            ws.cell(row=row, column=c).border = s["border"]
        row += 1

    # Print setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def main():
    parser = argparse.ArgumentParser(description="Generate restaurant proposal XLSX")
    parser.add_argument("--rate", type=float, default=300.0, help="Day rate in EUR (default: 300)")
    parser.add_argument("--tva", type=float, default=20.0, help="TVA percentage (default: 20)")
    args = parser.parse_args()

    wb = Workbook()
    build_proposal_sheet(wb, args.rate, args.tva)

    out = Path(__file__).resolve().parent.parent / "Proposition_Commerciale_Restaurant.xlsx"
    wb.save(out)
    print(f"Fichier genere: {out}")
    print(f"Taux: {args.rate} EUR/jour | TVA: {args.tva}%")


if __name__ == "__main__":
    main()
