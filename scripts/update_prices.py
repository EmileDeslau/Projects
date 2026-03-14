#!/usr/bin/env python3
"""Recalculate unit prices in Cahier_des_charges_resto.xlsx
Usage: python scripts/update_prices.py --rate 300
"""
import argparse
from openpyxl import load_workbook
from pathlib import Path

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--rate', type=float, default=300.0, help='day rate in EUR')
    parser.add_argument('--min', type=float, default=50.0, help='minimum unit price')
    parser.add_argument('--round', type=float, default=25.0, help='rounding step')
    args = parser.parse_args()
    p = Path('Cahier_des_charges_resto.xlsx')
    wb = load_workbook(p)
    ws = wb['RFP - Offre de prix']
    for r in range(2, ws.max_row+1):
        fid = ws.cell(r,1).value
        if not isinstance(fid, str):
            continue
        if fid.strip() == '':
            continue
        dur = ws.cell(r,5).value or 0
        try:
            price = float(dur) * float(args.rate)
            step = float(args.round)
            price = int(round(price/step)*step)
        except Exception:
            price = float(args.min)
        if price < float(args.min):
            price = float(args.min)
        ws.cell(r,6).value = price
    wb.save(p)

if __name__ == '__main__':
    main()